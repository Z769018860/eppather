import os
import re
import openpyxl
from openpyxl.styles import Font

# 定义函数：计算圈复杂度
def calculate_cyclomatic_complexity(code):
    keywords = [r'\bif\b', r'\belse if\b', r'\bfor\b', r'\bwhile\b', r'\bcase\b', r'\&\&', r'\|\|']
    complexity = 1  # 初始复杂度为 1
    for keyword in keywords:
        matches = re.findall(keyword, code)
        complexity += len(matches)
    return complexity

# 定义函数：计算代码行数
def count_code_lines(code):
    lines = code.split('\n')
    # 过滤掉空行和仅有注释的行
    valid_lines = [line for line in lines if line.strip() and not line.strip().startswith('//') and not re.match(r'^\s*/\*.*\*/\s*$', line)]
    return len(valid_lines)

# 定义函数：检查是否有数组定义
def contains_array(code):
    array_pattern = r'\b[a-zA-Z_][a-zA-Z0-9_]*\s*\[\s*[0-9]*\s*\]'
    return bool(re.search(array_pattern, code))

# 定义函数：检查是否有嵌套循环或条件语句
def contains_nested_structure(code):
    # 通过简单缩进逻辑检测嵌套结构
    nested_patterns = [
        r'\bfor\b.*\{[^}]*\bfor\b',
        r'\bwhile\b.*\{[^}]*\bwhile\b',
        r'\bif\b.*\{[^}]*\bif\b'
    ]
    for pattern in nested_patterns:
        if re.search(pattern, code, re.DOTALL):
            return True
    return False

# 定义函数：检查是否有位运算
def contains_bitwise_operations(code):
    # 匹配常见的位运算符
    bitwise_operators = [r'\&', r'\|', r'\^', r'<<', r'>>', r'~']
    for operator in bitwise_operators:
        if re.search(r'\b' + operator + r'\b', code):
            return True
    return False

# 定义函数：统计函数数量
def count_functions(code):
    # 匹配函数定义的正则表达式
    function_pattern = r'\b[a-zA-Z_][a-zA-Z0-9_]*\s+[a-zA-Z_][a-zA-Z0-9_]*\s*\([^)]*\)\s*\{'
    return len(re.findall(function_pattern, code))

# 定义函数：计算函数参数数量
def count_function_parameters(code):
    function_pattern = r'\b[a-zA-Z_][a-zA-Z0-9_]*\s+[a-zA-Z_][a-zA-Z0-9_]*\s*\(([^)]*)\)'
    matches = re.findall(function_pattern, code)
    param_counts = []
    for match in matches:
        params = match.split(',')
        param_counts.append(len([param for param in params if param.strip()]))  # 去掉空参数
    return param_counts

# 批量处理所有文件并保存到 Excel
def process_c_files_and_export_to_excel(folder_path, output_excel):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "C File Analysis"
    
    # 写入标题行
    headers = [
        "File Name", "Cyclomatic Complexity", "Code Lines", "Contains Array", 
        "Function Parameters", "Contains Nested Structure", "Contains Bitwise Operations", "Function Count"
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 遍历文件夹中的 C 文件
    row = 2
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.c'):  # 只处理 .c 文件
            file_path = os.path.join(folder_path, file_name)
            with open(file_path, 'r', encoding='utf-8') as f:
                code = f.read()
            
            # 分析文件
            complexity = calculate_cyclomatic_complexity(code)
            code_lines = count_code_lines(code)
            has_array = contains_array(code)
            nested_structure = contains_nested_structure(code)
            bitwise_operations = contains_bitwise_operations(code)
            function_count = count_functions(code)
            param_counts = count_function_parameters(code)
            
            # 写入数据到表格
            sheet.cell(row=row, column=1, value=file_name)
            sheet.cell(row=row, column=2, value=complexity)
            sheet.cell(row=row, column=3, value=code_lines)
            sheet.cell(row=row, column=4, value="Yes" if has_array else "No")
            sheet.cell(row=row, column=5, value=", ".join(map(str, param_counts)))
            sheet.cell(row=row, column=6, value="Yes" if nested_structure else "No")
            sheet.cell(row=row, column=7, value="Yes" if bitwise_operations else "No")
            sheet.cell(row=row, column=8, value=function_count)
            row += 1
    
    # 保存 Excel 文件
    workbook.save(output_excel)
    print(f"Analysis saved to {output_excel}")

# 主程序入口
if __name__ == "__main__":
    current_folder = os.getcwd()  # 当前文件夹
    output_file = os.path.join(current_folder, "c_file_analysis.xlsx")  # 输出 Excel 文件
    process_c_files_and_export_to_excel(current_folder, output_file)
